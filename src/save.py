import pandas as pd
from openpyxl import load_workbook
from io import BytesIO

def save(results):
    output_df = pd.DataFrame(results)

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        output_df.to_excel(writer, index=False, sheet_name='HR Profiles')

    output.seek(0)
    wb = load_workbook(output)
    ws = wb.active

    for row in range(2, ws.max_row + 1):
        link_cell = ws[f'C{row}']
        url = link_cell.value
        if url:
            link_cell.value = f'=HYPERLINK("{url}", "{url}")'

    final_output = BytesIO()
    wb.save(final_output)
    final_output.seek(0)

    return final_output
